import sys
import requests
import openpyxl
from openpyxl import Workbook
from datetime import datetime


filename = "iq_results.xlsx"
Contest = "algothon2021"

def read_participants(fname):

    excel_file = openpyxl.load_workbook(fname)

    act_sheet = excel_file.worksheets[0]

    sheet_data = []

    max_row_count = act_sheet.max_row

    #print("Max Row Count is : {}".format(str(max_row_count)))

    for x in range(2, max_row_count+1):

        sheet_row = []

        for y in range(1, 7):

            cell_dt = act_sheet.cell(row=x, column=y).value

            if(cell_dt == None and y==1):
                break

            if(y==3):

                cell_dt.upper()
                

            sheet_row.append(cell_dt)
        if(len(sheet_row) == 0):
            continue
        sheet_data.append(sheet_row)

    
    return sheet_data


def get_content_users(contest):

    users = []
    total = 0 

    try:
        while(True):
            

            headers_list = {
                "Accept": "*/*", "Accept-Encoding" : "gzip, deflate, br" , "Connection" : "keep-alive",
                "User-Agent" : "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.190 Safari/537.36"}

            res = requests.get("https://www.hackerrank.com/rest/contests/{}/leaderboard?offset={}&limit=100".format(contest,total),headers = headers_list)
            
            if(res.status_code == 200):

                res_wrapper = res.json()
                model_wrap = res_wrapper["models"]

                if(len(model_wrap) == 0):

                    return users

                for user_wrap in model_wrap:

                    users.append({"id" : user_wrap["hacker"].upper(), "score" : user_wrap["score"] })

                total+=100



    except Exception as e:
        print(e)

        print("User list Capturing faild")

        return e

def create_excel(leaderboard,prefix):

    now = datetime.now()
    current_time = now.strftime("%H-%M-%S ")


    sheet = Workbook()
    active_sheet = sheet.active

    for team in leaderboard:

        active_sheet.append(team)


    sheet.save(current_time + prefix + " leaderboard.xlsx")
    
if __name__ == "__main__":

    
    iq_list = read_participants(filename)
    print("Excel file reading completed")

    hacker_rank_list = get_content_users(Contest)
    print("Hacker Rank data fetched succussfully !")

    final_marks = []

    count = 0

    print("Marks checking now ..")

    for h_user in hacker_rank_list:

        found = False

        for iq_user in iq_list:

            if(h_user["id"] == iq_user[2]):

                found = True

                final_marks.append([h_user["id"] , iq_user[1] , iq_user[3] ,  iq_user[5] + h_user["score"]])

        if(not found):

            final_marks.append([ h_user["id"] , "" , "" ,  h_user["score"]])



    for iq_user in iq_list:

        found = False

        for h_user in hacker_rank_list:


            if(iq_user[2] == h_user["id"]):

                found = True
                break

        if(not found):
            
            final_marks.append([ iq_user[2] , iq_user[1] , iq_user[3] ,  iq_user[5]])

            


    print("Sorting...")

    final_marks = sorted(final_marks, reverse=True , key=lambda x: x[3])

    sliit_emails_only = []

    for x in final_marks:

        if("IT" == x[0][0:2] or "EN" == x[0][0:2]):

            sliit_emails_only.append(x)



    create_excel(final_marks,"All")
    create_excel(sliit_emails_only,"SLIIT Only")

    print("Completed !")
